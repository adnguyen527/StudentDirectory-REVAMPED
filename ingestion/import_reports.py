"""
Import anonymized data from Excel files into MongoDB.

DWP rows are parsed at import time -- compound string fields are split into
discrete typed fields before insertion.

Imports are idempotent: every row carries a row_hash content fingerprint, writes are
upserts keyed on it, and a unique index enforces it. Re-importing a file that has
already been loaded reports its rows as "unchanged" instead of duplicating them.

Note that a row edited at the source produces a different hash, so it is imported as a
new document rather than replacing the original. Correcting an already-imported row is
a separate operation from re-importing a file.

Run ingestion/backfill_row_hash.py once before the first import, to hash documents
loaded before this mechanism existed.
"""

import hashlib
import re
from pathlib import Path
from datetime import datetime
from bson import json_util
from pymongo import MongoClient, ASCENDING
from pymongo.errors import BulkWriteError
from mongo_url import uri, db_name
import openpyxl


# ── DWP parsing helpers ───────────────────────────────────────────────────────

def _split(value):
    """'Key: Value;  Key: Value' -> dict"""
    if not value:
        return {}
    result = {}
    for part in str(value).split(';  '):
        part = part.strip()
        if ': ' in part:
            key, val = part.split(': ', 1)
            result[key.strip()] = val.strip()
    return result

def _int(value):
    if value is None or str(value).strip() == 'None':
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None

def _bool(value):
    if value is None:
        return None
    return str(value).strip().lower() == 'yes'

def _none(value):
    return None if (value is None or str(value).strip() == 'None') else value

def _to_snake(s):
    return re.sub(r'[^a-z0-9]+', '_', str(s).lower()).strip('_')

def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip(), '%m/%d/%Y')
    except ValueError:
        return None


def _parse_clock(value):
    """'3:58 PM' -> datetime.time, or None if the cell holds nothing usable."""
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == 'None':
        return None
    try:
        return datetime.strptime(text.upper(), '%I:%M %p').time()
    except ValueError:
        return None


def combine_session_time(session_date, value):
    """The session date plus a clock reading -> one datetime.

    A time with no date attached cannot be sorted, ranged or subtracted without knowing
    which day it belongs to, so the two halves are joined once here rather than rejoined
    by every consumer. Returns None if either half is missing: a time on no date is not
    a moment.
    """
    if isinstance(value, datetime):
        return value
    clock = _parse_clock(value)
    if clock is None or session_date is None:
        return None
    return datetime.combine(session_date.date(), clock)


def _parse_finalized_date(value):
    """' 01/02/2025 \\n 3:59 PM' -> datetime(2025, 1, 2, 15, 59)

    The source packs the date and the time of finalization into one cell, separated by a
    newline and padded with spaces. Stored raw, it is a string nobody can range-query --
    'when was this report actually closed out' needs a real datetime.

    Whitespace is collapsed rather than split on '\\n' specifically, so a cell that uses
    a different separator still parses. A date with no time is accepted as midnight.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = ' '.join(str(value).split())
    if not text or text == 'None':
        return None
    for fmt in ('%m/%d/%Y %I:%M %p', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


# ONE TIME BACKFILL
# Anonymization artifacts. The name mapping assigned a person's name to rows whose
# instructor was empty in the source, so an unattributed session became one belonging to
# someone who does not exist -- 73 rows, and no row anywhere with an empty instructors
# list, which silently disabled the "no instructor named" guard in build_instructors.py.
# Names here are dropped at parse time so the emptiness survives into the database.
PLACEHOLDER_INSTRUCTORS = {'Elizabeth Griffin'}


def parse_instructors(value):
    """'Dana Reyes, Sam Ortiz' -> ['Dana Reyes', 'Sam Ortiz'], placeholders removed.

    A row left with no instructors is unattributed, which is a fact about the row and
    not a reason to drop it: the session still happened and still counts for the student.
    """
    if not value:
        return []
    names = [name.strip() for name in str(value).split(',')]
    return [n for n in names if n and n not in PLACEHOLDER_INSTRUCTORS]


def parse_session(value):
    p = _split(value)
    instructors_str = p.get('Instructors', '')
    return {
        'sessions_this_month': _int(p.get('Sessions This Month')),
        # _none(), like every other optional field here: the source writes the literal
        # string 'None' for a session with no recorded time, and it has to land as a
        # null. Without this, 'None' is stored as a value and every consumer has to know
        # to special-case it.
        'session_start':       _none(p.get('Session Start')),
        'session_end':         _none(p.get('Session End')),
        'instructors':         parse_instructors(instructors_str),
    }

def parse_general_information(value):
    p = _split(value)
    return {
        'session_page_goal': _int(p.get('Session Page Goal')),
        'pages_completed':   _int(p.get('Pages Completed')),
        'mathlete_score':    _int(p.get('Mathlete Score')),
        'last_punch_of_day': _bool(p.get('Last Punch of the Day'))
    }

def parse_digital_reward_system(value):
    p = _split(value)
    stars_str = p.get('Stars on current card', '0')
    if '/' in stars_str:
        left, right = stars_str.split('/', 1)
        stars_current, stars_max = _int(left.strip()), _int(right.strip())
    else:
        stars_current, stars_max = _int(stars_str), None

    session_stars_str = p.get('Session', '0 stars added')
    return {
        'card_level':          _none(p.get('Card level')),
        'stars_current':       stars_current,
        'stars_max':           stars_max,
        'session_stars_added': _int(session_stars_str.split(' ')[0])
    }

def parse_student_materials(value):
    p = _split(value)
    result = {
        'primary_deck_next_page':      _none(p.get('Start page for next session - Primary Deck')),
        'needs_primary_deck_update':   _bool(p.get('Needs Primary Deck update')),
        'secondary_deck_next_page':    _none(p.get('Start page for next session - Secondary Deck')),
        'needs_secondary_deck_update': _bool(p.get('Needs Secondary Deck update')),
        'internet_rating':             _none(p.get('Internet Rating'))
    }
    if 'Problem of the Week' in p:
        result['problem_of_the_week'] = _bool(p['Problem of the Week'])
    return result

def parse_schoolwork(value):
    p = _split(value)
    duration_str = p.get('Duration', '0 min')
    return {
        'schoolwork_start_time':  _none(p.get('Start time')),
        'schoolwork_duration_min': _int(duration_str.split(' ')[0]),
        'schoolwork_completed':   _bool(p.get('Completed')),
        'schoolwork_checked':     _bool(p.get('Checked')),
        'schoolwork_description': _none(p.get('Description'))
    }

def parse_lp_assignment(value):
    if not value:
        return []
    topics = []
    for entry in str(value).split(';  '):
        entry = entry.strip()
        if not entry:
            continue
        match = re.match(r'^([\w-]+)\s+\((.+)\):\s+(.+)$', entry)
        if match:
            topics.append({'id': match.group(1), 'name': match.group(2), 'status': match.group(3).strip()})
        else:
            topics.append({'raw': entry})
    return topics


COMPOUND_FIELDS = ['Session', 'General Information', 'Digital Reward System', 'Student Materials', 'LP Assignment', 'Schoolwork', 'Center']

def parse_center(value):
    """'Southlake, Mann Mathematics' -> {'centers': ['Southlake'],
                                         'center_orgs': ['Mann Mathematics']}

    The Center cell is '<location>, <organization>'. The organization is the operating
    brand on the date of the session, not a property of the student or the session:
    every location switched from 'Mann Mathematics' to 'Math Made Simple' on 2025-09-05,
    so a student attending either side of that date would otherwise appear to have
    attended two different centers. 1,438 rows carry a bare location and no organization.

    '@Home Classroom 1' (37 rows, Southlake, Aug-Dec 2024) sits in the organization
    position but names a room rather than a brand. It is kept as-is rather than special
    cased -- the position is what this function knows about.
    """
    if not value:
        return {'centers': [], 'center_orgs': []}

    centers, orgs = [], []
    for entry in str(value).split(';  '):
        entry = entry.strip()
        if not entry:
            continue
        # partition, not split: a location containing a comma would otherwise lose
        # everything past the first one.
        location, _, org = entry.partition(', ')
        location = location.strip()
        org = org.strip()
        if location and location not in centers:
            centers.append(location)
        if org and org not in orgs:
            orgs.append(org)
    return {'centers': centers, 'center_orgs': orgs}

def row_hash(doc):
    """Content fingerprint of a document, used as its idempotency key.

    A composite of (account_id, student_name, date, session_start) is NOT unique --
    four session pairs in the current data share one while being genuinely different
    records (different instructors, different notes). Hashing the whole document
    avoids collapsing those. All 29,382 existing documents hash distinctly.

    json_util handles BSON types (datetime, ObjectId) deterministically; sort_keys
    makes the digest independent of field insertion order.
    """
    body = {k: v for k, v in doc.items() if k not in ('_id', 'row_hash')}
    return hashlib.sha1(json_util.dumps(body, sort_keys=True).encode()).hexdigest()


def _is_row_hash_conflict(write_error):
    """True if a duplicate-key error came from the row_hash index rather than another.

    keyPattern is present on modern servers; the message is the fallback.
    """
    key_pattern = write_error.get('keyPattern')
    if key_pattern is not None:
        return 'row_hash' in key_pattern
    return 'row_hash' in write_error.get('errmsg', '')


def transform_dwp_row(row):
    doc = {_to_snake(k): v for k, v in row.items() if k not in COMPOUND_FIELDS}
    doc['date'] = _parse_date(row.get('Date'))
    # Read back off the snake-cased doc rather than the raw row, so this does not depend
    # on the source's column heading staying spelled the way it is today.
    doc['finalized_date'] = _parse_finalized_date(doc.get('finalized_date'))
    doc.update(parse_session(row.get('Session')))
    # parse_session sees only the Session cell, so it cannot know the date. Join the two
    # halves here, where both are in hand.
    doc['session_start'] = combine_session_time(doc['date'], doc.get('session_start'))
    doc['session_end'] = combine_session_time(doc['date'], doc.get('session_end'))
    doc.update(parse_general_information(row.get('General Information')))
    doc.update(parse_digital_reward_system(row.get('Digital Reward System')))
    doc.update(parse_student_materials(row.get('Student Materials')))
    doc.update(parse_schoolwork(row.get('Schoolwork')))
    doc.update(parse_center(row.get('Center')))
    doc['topics'] = parse_lp_assignment(row.get('LP Assignment'))
    doc['finalized'] = is_finalized(doc)
    doc['row_hash'] = row_hash(doc)
    return doc


def is_finalized(doc):
    """Was this session's report actually completed?

    Keyed on pages_completed, not finalized_date. 1,068 rows have no page count; 996 of
    them also have no finalized_date, but 547 other rows carry a finalized_date with a
    real page count missing from neither -- and 72 rows (all December 2024) are the
    reverse, finalized with no pages. A page count is the signal that survives both.

    An unfinalized row is still a session that happened: 968 of the 996 are the only
    record of that student-day, and all 996 name an instructor. They belong in
    attendance and out of any pages-per-session rate, which is what this flag is for.
    """
    return doc.get('pages_completed') is not None


# ── Importer ──────────────────────────────────────────────────────────────────

class DataImporter:

    def __init__(self):
        self.client = MongoClient(uri)
        self.db = self.client[db_name]
        self.stats = {
            'dwp_reports': 0,
            'attendance_reports': 0,
            'enrollment_reports': 0,
            'student_reports': 0,
            'birthday_reports': 0,
            'total_files': 0,
            'total_documents': 0,
            'already_present': 0,
            'repeated_in_file': 0,
            'errors': 0
        }

    def _read_excel(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        docs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            docs.append({h: v for h, v in zip(headers, row) if h})
        return docs

    def _collection_name(self, filename):
        name = filename.lower()
        if 'digital workout plan' in name or 'dwp' in name:
            return 'dwp_reports'
        if 'attendance' in name:
            return 'attendance_reports'
        if 'enrollment' in name:
            return 'enrollment_reports'
        if 'birthday' in name:
            return 'birthday_reports'
        if 'student' in name and 'report' in name:
            return 'student_reports'
        return None

    def import_file(self, path):
        filename = Path(path).name
        collection_name = self._collection_name(filename)
        if not collection_name:
            print(f"    [warn] Unknown file type, skipping: {filename}")
            return

        rows = self._read_excel(path)
        if not rows:
            print(f"    [warn] No data: {filename}")
            return

        if collection_name == 'dwp_reports':
            rows = [transform_dwp_row(row) for row in rows]
        else:
            for row in rows:
                row['row_hash'] = row_hash(row)

        try:
            inserted, already, repeated = self._upsert(collection_name, rows)
            note = f", {repeated} repeated in file" if repeated else ""
            print(f"    [ok] {inserted} new, {already} already present{note} "
                  f"-> {collection_name}")
            self.stats[collection_name] += inserted
            self.stats['total_documents'] += inserted
            self.stats['already_present'] += already
            self.stats['repeated_in_file'] += repeated
        except Exception as e:
            print(f"    [!!] Insert error: {e}")
            self.stats['errors'] += 1

    def _upsert(self, collection_name, docs):
        """Write only rows whose row_hash is not already stored.

        There is deliberately no update path. row_hash covers the whole document, so a
        hash that already exists means a byte-identical document is already stored and
        rewriting it would change nothing. A row edited at the source hashes
        differently and is inserted as a new document.

        Returns (inserted, already_present, repeated_within_file).
        """
        collection = self.db[collection_name]
        self._ensure_indexes(collection_name)

        # Collapse rows repeated inside a single file before touching the database.
        by_hash = {d['row_hash']: d for d in docs}
        repeated = len(docs) - len(by_hash)

        hashes = list(by_hash)
        already = set()
        for i in range(0, len(hashes), 1000):
            chunk = hashes[i:i + 1000]
            already |= set(collection.distinct('row_hash', {'row_hash': {'$in': chunk}}))

        fresh = [d for h, d in by_hash.items() if h not in already]
        inserted = 0
        if fresh:
            try:
                inserted = len(collection.insert_many(fresh, ordered=False).inserted_ids)
            except BulkWriteError as e:
                # A row_hash duplicate here means another writer inserted the same row
                # between the read above and this write -- harmless, the row is stored.
                # A duplicate on any other index is a real problem and must not be
                # swallowed, so check which index actually collided.
                unexpected = [w for w in e.details['writeErrors']
                              if w['code'] != 11000 or not _is_row_hash_conflict(w)]
                if unexpected:
                    raise
                inserted = e.details['nInserted']

        return inserted, len(by_hash) - inserted, repeated

    def _ensure_indexes(self, collection_name):
        collection = self.db[collection_name]
        # Unique on row_hash is what actually enforces idempotency -- a second import
        # of the same row matches an existing document instead of inserting beside it.
        collection.create_index([('row_hash', ASCENDING)], unique=True)
        if collection_name == 'dwp_reports':
            collection.create_index([('date', ASCENDING)])
            collection.create_index([('account_id', ASCENDING)])

    def import_all(self, directory='anonymized_data'):
        print(f"\n{'='*60}\nIMPORTING FROM {directory}\n{'='*60}")
        files = sorted(Path(directory).glob('**/*.xlsx'))
        if not files:
            print(f"[!!] No Excel files found in {directory}")
            return False

        self.stats['total_files'] = len(files)
        print(f"[ok] Found {len(files)} files\n")

        for i, f in enumerate(files, 1):
            print(f"[{i}/{len(files)}] {f.name}")
            try:
                self.import_file(str(f))
            except Exception as e:
                print(f"    [!!] {e}")
                self.stats['errors'] += 1

        return True

    def print_stats(self):
        print(f"\n{'='*60}\nIMPORT COMPLETE\n{'='*60}")
        print(f"Files:           {self.stats['total_files']}")
        print(f"New documents:   {self.stats['total_documents']}")
        print(f"Already present: {self.stats['already_present']}  (skipped -- re-import is a no-op)")
        if self.stats['repeated_in_file']:
            print(f"Repeated in file: {self.stats['repeated_in_file']}")
        print(f"Errors:          {self.stats['errors']}")
        print(f"\nBy collection:")
        for name in ['dwp_reports', 'attendance_reports', 'enrollment_reports', 'student_reports', 'birthday_reports']:
            print(f"  {name}: {self.stats[name]}")

    def verify(self):
        print(f"\n{'='*60}\nVERIFICATION\n{'='*60}")
        for name in ['dwp_reports', 'attendance_reports', 'enrollment_reports', 'student_reports', 'birthday_reports']:
            count = self.db[name].count_documents({})
            print(f"  {name}: {count}")

    def close(self):
        self.client.close()


def main():
    importer = DataImporter()
    try:
        importer.db.command('ping')
        print("[ok] Connected to MongoDB")
        if importer.import_all('anonymized_data'):
            importer.print_stats()
            importer.verify()
    except Exception as e:
        print(f"[!!] {e}")
    finally:
        importer.close()


if __name__ == '__main__':
    main()
